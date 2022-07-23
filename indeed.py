import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import time

start_time = time.time()

wb = Workbook()
ws=wb.active

titleArray = []
companyArray = []
locationArray = []
linkArray = []
salaryArray = []


def page(pageNo, TitleSearch, LocationSearch):
    URL = "https://uk.indeed.com/jobs?q="+TitleSearch+"%20&l="+LocationSearch+"%20&start={}".format(pageNo)
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, "html.parser")
    jobs = soup.find_all("td", class_="resultContent")
    return jobs
    

def results(jobs, jobWord):
    mainURL = "https://uk.indeed.com"
    
    for job in jobs: 
        try:
            title = job.find("a", class_="jcs-JobTitle").text
        except:
            title = "No title found"
        try:
            company = job.find("span", class_="companyName").text
        except:
            company = "No company found"
        try:
            location = job.find("div", class_="companyLocation").text
        except:
            location = "No location found"
        try:
            salary = job.find("div", class_="metadata salary-snippet-container").text
        except:
            salary = "No salary found"
        try:         
            link = job.find("a").get("href")
        except:
            link = "No link found"
            
        if title.find(jobWord) or title.find(jobWord.lower()) != -1:
                titleArray.append(title)
                companyArray.append(company)
                locationArray.append(location)
                salaryArray.append(salary)
                linkArray.append(mainURL + link)
       
                
    ws.cell(row=1, column=1, value="Title")
    ws.cell(row=1, column=2, value="Company")
    ws.cell(row=1, column=3, value="Location")
    ws.cell(row=1, column=4, value="Salary")
    ws.cell(row=1, column=5, value="Apply Here")
        
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)
    ws["D1"].font = Font(bold=True)
    ws["E1"].font = Font(bold=True)
    
    length = len(titleArray)
    
    i = 2

    for j in range(0,length): 
        ws.cell(row=i, column=1, value=titleArray[j])
        ws.cell(row=i, column=2, value=companyArray[j])
        ws.cell(row=i, column=3, value=locationArray[j])
        ws.cell(row=i, column=4, value=salaryArray[j])
        ws.cell(row=i, column=5).hyperlink = linkArray[j]
        i+=1
        
    wb.save("jobs.xlsx")
        
    
    
def main():
    for page_order in range(0, 40, 10):
        extract = page(page_order, "Software Developer", "London")
        publish = results(extract, "Developer")
    
    
    # https://uk.indeed.com/jobs?q=Software%20Developer&l=London&start=10

main()


print(time.time() - start_time)